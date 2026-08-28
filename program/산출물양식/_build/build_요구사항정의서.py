# -*- coding: utf-8 -*-
"""요구사항 정의서 양식을 채운다.

양식은 A5:A40 구간에 병합 그룹이 미리 잡혀 있다(대분류 5개, 중분류 15개, 총 36행).
그 병합 모양을 그대로 지키고 값만 넣는다.
"""
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Font

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

HERE = os.path.dirname(os.path.abspath(__file__))
FORMS = os.path.dirname(HERE)
SRC = os.path.join(FORMS, "[모델배포]_요구사항 정의서_양식.xlsx")
OUT = os.path.join(FORMS, "[모델배포]_요구사항 정의서_A-COPilot.xlsx")

# (분류, 대분류, [(중분류, [(요구사항ID, 기능설명, 추가설명, 비고), ...]), ...])
GROUPS = [
    ("기능", "케이스 런타임 (Core 1)", [
        ("케이스 생명주기", [
            ("REQ-001", "외부 요청을 고객 케이스로 만들고 new 상태로 둔다.",
             "REST `POST /v1/cases`와 외부 AI의 `open_support_case` 두 경로로 들어온다.",
             "Core 1"),
            ("REQ-002", "케이스 상태 전이는 허용된 다음 상태 표를 벗어나면 거부한다.",
             "상태는 new·classifying·routing·running·waiting_input·waiting_approval·"
             "waiting_external·resuming·resolved·escalated·failed·cancelled의 12종이다.",
             "Core 1"),
            ("REQ-003", "모든 상태 전이를 case_events에 추가만 하고, 나중에 그대로 재생할 수 있어야 한다.",
             "기존 행을 고치지 않는 append-only 방식이다. 감사와 재현의 근거가 된다.",
             "증적 DoD-03"),
        ]),
        ("인라인 분류", [
            ("REQ-004", "케이스 생성 경로에서 감성·의도·이슈를 함께 분류한다.",
             "선택 기능이 아니다. 의도는 billing·technical·other로 나누고 이슈는 코드와 심각도를 함께 저장한다.",
             "모델"),
            ("REQ-005", "분류에 실패하면 classification_failed를 남기고 escalated로 전환한다.",
             "실패를 조용히 넘기지 않는다. 사람이 받도록 명시적으로 올린다.",
             "모델"),
        ]),
        ("팀 라우팅", [
            ("REQ-006", "Agent Team Registry가 케이스 유형과 의도로 담당 팀을 찾는다.",
             "팀은 코드에 박히지 않고 선언으로 등록된다.", "Core 1"),
            ("REQ-007", "조건 하나에 팀이 정확히 하나만 나와야 한다. 0개나 2개면 no_team으로 처리한다.",
             "찾지 못하면 ROUTING_FAILED 이벤트를 남기고 escalated로 보낸다.", "Core 1"),
            ("REQ-008", "Core는 팀 내부의 그래프·프롬프트·검색 코드를 import하지 않는다.",
             "TeamManifest와 표준 계약만 사용한다. 팀 추가가 Core 수정으로 번지지 않게 하는 규칙이다.",
             "증적 DoD-22"),
        ]),
        ("컨텍스트 조합", [
            ("REQ-009", "Context Broker가 팀 권한 범위 안에서 근거를 읽어 ContextPack으로 묶는다.",
             "RAG 문서, DB 현재 상태, 케이스 이력, Memory를 골라 조합한다.", "Core 1"),
            ("REQ-010", "팀은 조회 도구를 직접 호출하지 않는다.",
             "직접 호출을 허용하면 팀마다 권한 범위가 달라져 통제가 무너진다.", "증적 DoD-22"),
            ("REQ-011", "근거가 부족하면 팀이 need_more_context 신호로 Controller에 다시 요청한다.",
             "팀이 스스로 자료를 더 가져오지 않고 Controller를 통해 요청한다.", "Core 1"),
        ]),
        ("공유 상태와 동시성", [
            ("REQ-012", "Shared State 갱신은 CAS(비교 후 교체) 방식으로만 허용한다.",
             "버전이 다르면 쓰기를 거부하고 다시 읽게 한다.", "동시성 요구는 세 항목을 함께 본다. "
             "여러 에이전트가 같은 케이스를 동시에 건드리는 상황이 전제다."),
            ("REQ-013", "충돌 시 재시도하며, 나중 쓰기가 앞선 쓰기를 조용히 덮어쓰지 않는다.",
             "덮어쓰기가 일어나면 앞선 판단의 근거가 사라진다.", None),
            ("REQ-014", "Controller의 재계획 횟수에 상한을 두고 초과하면 즉시 중단한다.",
             "에이전트끼리 서로를 계속 호출하는 무한 루프를 차단한다.", None),
        ]),
        ("이벤트 전달", [
            ("REQ-015", "Message Broker는 Task와 Event의 배달만 한다. 팀 선택이나 업무 판단은 하지 않는다.",
             "배달 계층이 판단을 겸하면 재시도와 중복 전달이 곧바로 업무 오류가 된다.", "Core 1"),
            ("REQ-016", "Outbox로 DB 기록과 메시지 발행을 한 트랜잭션으로 묶는다.",
             "기록은 됐는데 메시지가 안 나가는 상황을 막는다.", "증적 DoD-12"),
            ("REQ-017", "중복 전달과 재시도는 수신 측 규칙으로 흡수한다.",
             "같은 메시지를 두 번 받아도 결과가 달라지지 않아야 한다.", "Core 1"),
        ]),
    ]),
    ("기능", "액션 실행과 승인 (Core 2)", [
        ("액션 제안 검증", [
            ("REQ-018", "팀은 부수효과를 실행하지 않고 ActionProposal(작업 제안)만 반환한다.",
             "환불·발주 같은 실제 영향은 Core 2 한 곳에서만 일어난다.", "증적 DoD-24"),
            ("REQ-019", "ContextPack이나 DB에 없는 근거를 든 제안은 거부하고 사유를 기록한다.",
             "근거 정합률과 근거 초과율로 측정한다. 쓰기 권한을 여는 전제 조건이다.", "Core 2"),
        ]),
        ("승인", [
            ("REQ-020", "고위험 액션은 케이스를 waiting_approval로 두고 사람이 승인해야 실행한다.",
             "무엇이 고위험인지는 정책으로 정하고 코드에 박지 않는다.", "Core 2"),
            ("REQ-021", "승인·거부의 주체와 시각을 action_approvals에 기록한다.",
             "누가 결재했는지 남지 않으면 감사에서 책임을 물을 수 없다.", "Core 2"),
            ("REQ-022", "waiting_external에서의 재개는 resume_token 검증을 통과해야 한다.",
             "결제사·택배사 콜백을 가장한 위조 재개를 막는다.", "Core 2"),
        ]),
        ("멱등성과 감사", [
            ("REQ-023", "같은 요청이 10회 들어와도 실제 부수효과는 1회만 일어난다.",
             "재시도가 중복 실행이 되지 않게 멱등성 키로 막는다.", "증적 DoD-11"),
            ("REQ-024", "모든 액션 실행을 audit_logs에 남긴다.",
             "누가 무엇을 왜 했는지를 사후에 확인할 수 있어야 한다.", "Core 2"),
        ]),
    ]),
    ("인터페이스", "외부 AI 연동", [
        ("REST / MCP / A2A", [
            ("REQ-025", "REST/OpenAPI로 케이스 생성과 상태 조회를 제공한다.",
             "MVP는 5개 endpoint로 시작한다. 이 숫자는 상한이 아니라 평가 범위다.", "Core 2"),
            ("REQ-026", "MCP 도구로 외부 AI가 케이스를 열고 상태를 조회한다.",
             "지금 여는 것은 제한된 검증 범위의 문의와 조회다. 전면 Commerce Ops는 확장 범위다.",
             "Core 2"),
            ("REQ-027", "A2A Task 상태를 케이스 상태에 매핑하고 원격 위임 결과를 받는다.",
             "A2A는 에이전트끼리 작업을 주고받는 규약이다. Catalog & Verification이 이 경로를 쓴다.",
             "증적 DoD-14"),
        ]),
    ]),
    ("기능", "도메인 팩 (Agent Team)", [
        ("VOC & Store Manager", [
            ("REQ-028", "이상징후와 급증을 탐지하고, 원인 축을 판별해 다른 팀에 위임을 제안한다.",
             "하루 1회 규칙 기반 집계로 시작한다. 임베딩 클러스터링과 토픽 모델링은 쓰지 않는다.",
             "모델"),
        ]),
        ("Response Generation & Review", [
            ("REQ-029", "근거를 인용한 답변 초안을 만들고 자체 검수를 거친다.",
             "인용 없는 답변은 내보내지 않는다.", "증적 DoD-29"),
            ("REQ-030", "개인정보 노출, 정책 위반, 근거 없는 주장을 내보내기 전에 차단한다.",
             "검사 규칙은 config/guardrails.yaml에 두고 코드와 분리한다.", "모델"),
        ]),
        ("Commerce Ops Pack", [
            ("REQ-031", "Procurement + Order & Payment와 Fulfillment & Logistics가 판단과 제안을 담당한다.",
             "확장 범위다. 검증 쇼핑몰이 준비되는 대로 진행한다. 이번 기간 착수 범위가 아니다.", "확장 범위"),
            ("REQ-032", "Return & Refund는 Registry 계약과 Mock으로 두고, Catalog & Verification은 A2A Remote로 둔다.",
             "Mock은 계약만 고정하고 실제 실행은 하지 않는다는 뜻이다. 평가 배분은 0건이다.",
             "확장 범위"),
        ]),
    ]),
    ("기능", "운영 콘솔과 Composer", [
        ("운영 콘솔", [
            ("REQ-033", "케이스·이벤트·승인 대기·평가 결과를 한 화면에서 조회한다.",
             "대상 시스템이 떠 있지 않아도 콘솔 자체는 떠야 한다.", "검증·프론트"),
            ("REQ-034", "콘솔은 대상(final_project_cs)의 Python을 import하지 않고 인증된 HTTP만 호출한다.",
             "대상 파일을 직접 쓰지 않는다. 쓰기는 대상 프로세스 안에서 대상의 계약으로 검증된 뒤 실행된다.",
             "final_project_ui/CLAUDE.md §0.3"),
        ]),
        ("Composer (구성 관리)", [
            ("REQ-035", "등록된 모듈과 Team의 활성 상태를 UI에서 바꾼다.",
             "POST /composer/toggle. 이미 선언된 항목만 켜고 끈다. 새 항목은 만들지 않는다.",
             "구현됨. v2 apply와 병행"),
            ("REQ-036", "카탈로그에서 구현 종류를 골라 인스턴스를 생성·수정·삭제한다.",
             "GET /composer/catalog와 POST /composer/changes 계약. 선언형 Team 실행기 배포가 선행 조건이다.",
             "미구현. 계약 확정 단계"),
        ]),
    ]),
]

wb = openpyxl.load_workbook(SRC)
ws = wb.active

FONT = Font(name="맑은 고딕", size=10)
FONT_B = Font(name="맑은 고딕", size=10, bold=True)
WRAP = Alignment(vertical="center", wrap_text=True)
WRAP_C = Alignment(vertical="center", horizontal="center", wrap_text=True)

row = 5
for kind, major, mids in GROUPS:
    group_start = row
    ws.cell(group_start, 1).value = kind
    ws.cell(group_start, 1).alignment = WRAP_C
    ws.cell(group_start, 1).font = FONT
    ws.cell(group_start, 2).value = major
    ws.cell(group_start, 2).alignment = WRAP_C
    ws.cell(group_start, 2).font = FONT_B
    for mid, items in mids:
        ws.cell(row, 3).value = mid
        ws.cell(row, 3).alignment = WRAP_C
        ws.cell(row, 3).font = FONT
        for req_id, desc, extra, note in items:
            ws.cell(row, 4).value = req_id
            ws.cell(row, 4).alignment = WRAP_C
            ws.cell(row, 4).font = FONT
            ws.cell(row, 5).value = desc
            ws.cell(row, 5).alignment = WRAP
            ws.cell(row, 5).font = FONT
            ws.cell(row, 6).value = extra
            ws.cell(row, 6).alignment = WRAP
            ws.cell(row, 6).font = FONT
            if note is not None:                    # G16:G18은 병합돼 있어 첫 행에만 쓴다
                ws.cell(row, 7).value = note
                ws.cell(row, 7).alignment = WRAP
                ws.cell(row, 7).font = FONT
            row += 1

assert row == 41, "행 수가 양식 병합 구간과 맞지 않음: %d" % row

ws.column_dimensions["B"].width = 22
ws.column_dimensions["C"].width = 20
ws.column_dimensions["D"].width = 12
ws.column_dimensions["F"].width = 46
ws.column_dimensions["G"].width = 22

note_row = row + 2
ws.cell(note_row, 1).value = "근거"
ws.cell(note_row, 1).font = FONT_B
ws.cell(note_row, 2).value = (
    "요구사항 본문은 program/plan/A-COP_구현계획서_v8.md의 §3-A·§7·§8·§8-B·§9-E·§19를 근거로 한다. "
    "Composer 두 항목은 program/plan/A-COP_Composer_범위재검토.md와 A-COP_Composer_소유권_정정.md를 따른다."
)
ws.cell(note_row, 2).font = FONT
ws.cell(note_row + 1, 2).value = (
    "REQ-036은 아직 구현되지 않았다. 선언형 Team 실행기 배포 → 인스턴스 CRUD 계약 확정 → "
    "카탈로그 HTTP 조회 → UI 선택 생성 화면 순서로 진행한다."
)
ws.cell(note_row + 1, 2).font = FONT
ws.cell(note_row + 2, 2).value = (
    "증적 DoD 번호는 final_project_cs/docs/evidence/ 아래 같은 번호의 문서를 가리킨다."
)
ws.cell(note_row + 2, 2).font = FONT

try:
    wb.save(OUT)
except PermissionError:
    # 엑셀로 열어 둔 파일은 덮어쓸 수 없다. 옆에 새 이름으로 저장하고 알린다.
    base, ext = os.path.splitext(OUT)
    OUT = base + "_새버전" + ext
    print("  원본이 열려 있어 새 이름으로 저장한다:", os.path.basename(OUT))
    wb.save(OUT)
print("저장:", OUT, "| 요구사항", sum(len(i) for _, _, m in GROUPS for _, i in m), "건")
